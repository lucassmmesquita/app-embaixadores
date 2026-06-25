"""
═══════════════════════════════════════════════════════════════
  Admin Module — Router
  Administrative endpoints for campaign management.
  PRD §7.1: Full admin panel functionality
  PRD §7.2: RBAC with role-specific access
═══════════════════════════════════════════════════════════════
"""

import csv
import io
import uuid
from datetime import datetime, timedelta, timezone
from typing import Annotated

from fastapi import APIRouter, Depends, Query, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import Integer, func, select, update as sql_update
from sqlalchemy.ext.asyncio import AsyncSession

from src.core.database import get_db
from src.modules.admin_auth.dependencies import get_current_admin_user
from src.modules.admin_auth.models import AdminUser
from src.modules.content.models import Content
from src.modules.content.schemas import ContentCreate, ContentUpdate
from src.modules.content.service import ContentService
from src.modules.events.models import Event, EventParticipant
from src.modules.events.schemas import EventCreate, EventResponse, EventUpdate
from src.modules.events.service import EventService
from src.modules.gamification.engine import GamificationEngine
from src.modules.gamification.models import Activity, Badge
from src.modules.gamification.schemas import BadgeCreate, BadgeResponse, BadgeUpdate
from src.modules.missions.models import Mission, UserMission
from src.modules.missions.schemas import MissionCreate, MissionResponse, MissionUpdate
from src.modules.missions.service import MissionService
from src.modules.notifications.models import Notification
from src.modules.notifications.service import NotificationService
from src.modules.notifications.system_config import SystemNotificationService
from src.modules.users.models import Level, Profile
from src.modules.users.schemas import LevelResponse, LevelUpdate, ProfileResponse
from src.modules.users.service import UserService
from src.shared.audit import AuditLog, log_audit
from src.shared.exceptions import BadRequestException, NotFoundException
from src.shared.pagination import PaginationParams

router = APIRouter()


# ═══ DASHBOARD ═══
@router.get("/dashboard/stats")
async def get_dashboard_stats(
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Get campaign dashboard statistics."""
    # Total users
    total_users = (await db.execute(
        select(func.count(Profile.id)).where(Profile.is_active.is_(True))
    )).scalar() or 0

    # New users this week (last 7 days)
    week_ago = datetime.now(timezone.utc) - timedelta(days=7)
    new_users_week = (await db.execute(
        select(func.count(Profile.id)).where(Profile.created_at >= week_ago)
    )).scalar() or 0

    # Total points awarded
    total_points = (await db.execute(
        select(func.sum(Activity.points_awarded))
    )).scalar() or 0

    # Total missions completed
    missions_completed = (await db.execute(
        select(func.count(UserMission.id)).where(UserMission.status == "completed")
    )).scalar() or 0

    # Total events
    total_events = (await db.execute(
        select(func.count(Event.id)).where(Event.is_active.is_(True))
    )).scalar() or 0

    # Active missions
    active_missions = (await db.execute(
        select(func.count(Mission.id)).where(Mission.is_active.is_(True))
    )).scalar() or 0

    # Users per level
    users_per_level = []
    levels_result = await db.execute(
        select(Level).order_by(Level.order_index)
    )
    for level in levels_result.scalars().all():
        count = (await db.execute(
            select(func.count(Profile.id)).where(Profile.current_level_id == level.id)
        )).scalar() or 0
        users_per_level.append({
            "level_name": level.name,
            "level_color": level.color,
            "count": count,
        })

    # Pending level approvals (PRD §3.2)
    pending_approvals = (await db.execute(
        select(func.count(Profile.id)).where(Profile.level_pending_approval.is_(True))
    )).scalar() or 0

    # Pending mission verifications
    pending_verifications = (await db.execute(
        select(func.count(UserMission.id)).where(UserMission.status == "submitted")
    )).scalar() or 0

    # Recent activities
    recent = await db.execute(
        select(Activity).order_by(Activity.created_at.desc()).limit(10)
    )

    return {
        "total_users": total_users,
        "new_users_week": new_users_week,
        "total_points_awarded": total_points,
        "missions_completed": missions_completed,
        "total_events": total_events,
        "active_missions": active_missions,
        "users_per_level": users_per_level,
        "pending_approvals": pending_approvals,
        "pending_verifications": pending_verifications,
        "recent_activities": list(recent.scalars().all()),
    }


# ═══ DASHBOARD EXPORT ═══
@router.get("/dashboard/export-excel")
async def export_engagement_excel(
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Export platform engagement data as single-sheet Excel report."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Engajamento"
    now = datetime.now()
    now_br = now.strftime("%d/%m/%Y %H:%M")

    # ═══ Style definitions ═══
    title_font = Font(name="Calibri", bold=True, size=16, color="1A202C")
    subtitle_font = Font(name="Calibri", size=11, color="718096", italic=True)
    section_font = Font(name="Calibri", bold=True, size=13, color="2D3748")
    section_fill = PatternFill(start_color="EDF2F7", end_color="EDF2F7", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    metric_label_font = Font(name="Calibri", size=11, color="4A5568")
    metric_value_font = Font(name="Calibri", bold=True, size=11, color="1A202C")
    border_bottom = Border(bottom=Side(style="thin", color="E2E8F0"))

    def write_section_title(r: int, title: str) -> int:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        cell = ws.cell(row=r, column=1, value=title)
        cell.font = section_font
        cell.fill = section_fill
        for col in range(1, 7):
            ws.cell(row=r, column=col).fill = section_fill
        return r + 1

    def write_metric(r: int, label: str, value) -> int:
        ws.cell(row=r, column=1, value=label).font = metric_label_font
        ws.cell(row=r, column=1).border = border_bottom
        ws.cell(row=r, column=2, value=value).font = metric_value_font
        ws.cell(row=r, column=2).border = border_bottom
        return r + 1

    def write_table_header(r: int, headers: list[str]) -> int:
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        return r + 1

    # ═══ TÍTULO ═══
    ws.merge_cells("A1:F1")
    ws.cell(row=1, column=1, value="Relatório de Engajamento — Rede de Embaixadores").font = title_font
    ws.cell(row=2, column=1, value=f"Data de extração: {now_br}").font = subtitle_font
    ws.merge_cells("A2:F2")
    ws.row_dimensions[1].height = 30
    r = 4

    # ═══ SEÇÃO 1: RESUMO GERAL ═══
    total_users = (await db.execute(
        select(func.count(Profile.id)).where(Profile.is_active.is_(True))
    )).scalar() or 0

    inactive_users = (await db.execute(
        select(func.count(Profile.id)).where(Profile.is_active.is_(False))
    )).scalar() or 0

    week_ago = now - timedelta(days=7)
    new_users_week = (await db.execute(
        select(func.count(Profile.id)).where(Profile.created_at >= week_ago)
    )).scalar() or 0

    month_ago = now - timedelta(days=30)
    new_users_month = (await db.execute(
        select(func.count(Profile.id)).where(Profile.created_at >= month_ago)
    )).scalar() or 0

    total_points = (await db.execute(
        select(func.sum(Activity.points_awarded))
    )).scalar() or 0

    r = write_section_title(r, "RESUMO GERAL")
    r = write_metric(r, "Total de Usuários Ativos", total_users)
    r = write_metric(r, "Usuários Inativos/Suspensos", inactive_users)
    r = write_metric(r, "Novos Usuários (últimos 7 dias)", new_users_week)
    r = write_metric(r, "Novos Usuários (últimos 30 dias)", new_users_month)
    r = write_metric(r, "Total de Pontos Distribuídos", total_points)
    r += 1

    # ═══ SEÇÃO 2: USUÁRIOS POR NÍVEL ═══
    r = write_section_title(r, "USUÁRIOS POR NÍVEL")
    r = write_table_header(r, ["Nível", "Usuários", "% do Total"])

    levels_result = await db.execute(select(Level).order_by(Level.order_index))
    for level in levels_result.scalars().all():
        count = (await db.execute(
            select(func.count(Profile.id)).where(Profile.current_level_id == level.id)
        )).scalar() or 0
        pct = f"{(count / total_users * 100):.1f}%" if total_users > 0 else "0%"
        ws.cell(row=r, column=1, value=level.name).border = border_bottom
        ws.cell(row=r, column=2, value=count).border = border_bottom
        ws.cell(row=r, column=3, value=pct).border = border_bottom
        r += 1
    r += 1

    # ═══ SEÇÃO 3: MISSÕES ═══
    total_missions = (await db.execute(select(func.count(Mission.id)))).scalar() or 0
    active_missions = (await db.execute(
        select(func.count(Mission.id)).where(Mission.is_active.is_(True))
    )).scalar() or 0
    users_started_missions = (await db.execute(
        select(func.count(func.distinct(UserMission.user_id)))
    )).scalar() or 0
    total_completions = (await db.execute(
        select(func.count(UserMission.id)).where(UserMission.status == "completed")
    )).scalar() or 0
    users_completed = (await db.execute(
        select(func.count(func.distinct(UserMission.user_id))).where(UserMission.status == "completed")
    )).scalar() or 0
    total_in_progress = (await db.execute(
        select(func.count(UserMission.id)).where(UserMission.status == "in_progress")
    )).scalar() or 0

    r = write_section_title(r, "MISSÕES")
    r = write_metric(r, "Total de Missões Cadastradas", total_missions)
    r = write_metric(r, "Missões Ativas", active_missions)
    r = write_metric(r, "Usuários que Iniciaram Missões", users_started_missions)
    r = write_metric(r, "Usuários que Concluíram ao menos 1", users_completed)
    r = write_metric(r, "Total de Conclusões", total_completions)
    r = write_metric(r, "Em Progresso", total_in_progress)
    denom = total_completions + total_in_progress
    comp_rate = f"{(total_completions / denom * 100):.1f}%" if denom > 0 else "0%"
    r = write_metric(r, "Taxa de Conclusão", comp_rate)
    r += 1

    # Detalhamento por missão
    r = write_table_header(r, ["Missão", "Iniciaram", "Concluíram", "Taxa"])
    missions_result = await db.execute(
        select(Mission).where(Mission.is_active.is_(True)).order_by(Mission.created_at.desc())
    )
    for mission in missions_result.scalars().all():
        m_started = (await db.execute(
            select(func.count(UserMission.id)).where(UserMission.mission_id == mission.id)
        )).scalar() or 0
        m_completed = (await db.execute(
            select(func.count(UserMission.id)).where(
                UserMission.mission_id == mission.id, UserMission.status == "completed"
            )
        )).scalar() or 0
        m_rate = f"{(m_completed / m_started * 100):.0f}%" if m_started > 0 else "—"
        ws.cell(row=r, column=1, value=mission.title).border = border_bottom
        ws.cell(row=r, column=2, value=m_started).border = border_bottom
        ws.cell(row=r, column=3, value=m_completed).border = border_bottom
        ws.cell(row=r, column=4, value=m_rate).border = border_bottom
        r += 1
    r += 1

    # ═══ SEÇÃO 4: EVENTOS ═══
    total_events_count = (await db.execute(select(func.count(Event.id)))).scalar() or 0
    active_events = (await db.execute(
        select(func.count(Event.id)).where(Event.is_active.is_(True))
    )).scalar() or 0
    total_registrations = (await db.execute(
        select(func.count(EventParticipant.id))
    )).scalar() or 0
    total_checkins = (await db.execute(
        select(func.count(EventParticipant.id)).where(EventParticipant.check_in_at.isnot(None))
    )).scalar() or 0
    unique_event_users = (await db.execute(
        select(func.count(func.distinct(EventParticipant.user_id)))
    )).scalar() or 0

    r = write_section_title(r, "EVENTOS")
    r = write_metric(r, "Total de Eventos", total_events_count)
    r = write_metric(r, "Eventos Ativos", active_events)
    r = write_metric(r, "Total de Inscrições", total_registrations)
    r = write_metric(r, "Total de Check-ins", total_checkins)
    r = write_metric(r, "Usuários Únicos em Eventos", unique_event_users)
    ck_rate = f"{(total_checkins / total_registrations * 100):.1f}%" if total_registrations > 0 else "0%"
    r = write_metric(r, "Taxa de Presença (check-in/inscrição)", ck_rate)
    r += 1

    # Detalhamento por evento
    r = write_table_header(r, ["Evento", "Data", "Local", "Inscritos", "Presentes", "Taxa Presença"])
    events_result = await db.execute(select(Event).order_by(Event.start_datetime.desc()))
    for event in events_result.scalars().all():
        e_reg = (await db.execute(
            select(func.count(EventParticipant.id)).where(EventParticipant.event_id == event.id)
        )).scalar() or 0
        e_ck = (await db.execute(
            select(func.count(EventParticipant.id)).where(
                EventParticipant.event_id == event.id, EventParticipant.check_in_at.isnot(None)
            )
        )).scalar() or 0
        e_rate = f"{(e_ck / e_reg * 100):.0f}%" if e_reg > 0 else "—"
        ws.cell(row=r, column=1, value=event.title).border = border_bottom
        ws.cell(row=r, column=2, value=event.start_datetime.strftime("%d/%m/%Y") if event.start_datetime else "—").border = border_bottom
        ws.cell(row=r, column=3, value=event.location_name or event.city or "Online").border = border_bottom
        ws.cell(row=r, column=4, value=e_reg).border = border_bottom
        ws.cell(row=r, column=5, value=e_ck).border = border_bottom
        ws.cell(row=r, column=6, value=e_rate).border = border_bottom
        r += 1
    r += 1

    # ═══ SEÇÃO 5: CONTEÚDO ═══
    total_content = (await db.execute(select(func.count(Content.id)))).scalar() or 0
    active_content = (await db.execute(
        select(func.count(Content.id)).where(Content.is_active.is_(True))
    )).scalar() or 0
    total_shares = (await db.execute(select(func.sum(Content.total_shares)))).scalar() or 0

    from src.modules.content.models import ContentShare
    unique_sharers = (await db.execute(
        select(func.count(func.distinct(ContentShare.user_id)))
    )).scalar() or 0

    r = write_section_title(r, "CONTEÚDO")
    r = write_metric(r, "Total de Conteúdos", total_content)
    r = write_metric(r, "Conteúdos Ativos", active_content)
    r = write_metric(r, "Total de Compartilhamentos", total_shares)
    r = write_metric(r, "Usuários que Compartilharam", unique_sharers)
    avg_sh = f"{(total_shares / active_content):.1f}" if active_content > 0 and total_shares else "0"
    r = write_metric(r, "Média de Shares por Conteúdo", avg_sh)
    r += 1

    # Detalhamento por conteúdo
    r = write_table_header(r, ["Conteúdo", "Tipo", "Compartilhamentos"])
    content_result = await db.execute(
        select(Content).where(Content.is_active.is_(True)).order_by(Content.total_shares.desc())
    )
    for content in content_result.scalars().all():
        ws.cell(row=r, column=1, value=content.title).border = border_bottom
        ws.cell(row=r, column=2, value=content.content_type).border = border_bottom
        ws.cell(row=r, column=3, value=content.total_shares).border = border_bottom
        r += 1

    # ═══ Auto-adjust column widths ═══
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 55)

    # ═══ Generate file ═══
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"embaixadores_engajamento_{now.strftime('%d_%m_%Y')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )




# ═══ LEVELS MANAGEMENT (PRD §3.1) ═══
@router.get("/levels")
async def list_levels(
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Admin: List all gamification levels ordered by progression."""
    result = await db.execute(select(Level).order_by(Level.order_index))
    levels = result.scalars().all()
    return [LevelResponse.model_validate(level) for level in levels]


@router.put("/levels/{level_id}")
async def update_level(
    level_id: uuid.UUID,
    data: LevelUpdate,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    request: Request,
):
    """
    Admin: Update a level's configuration (PRD §3.1).
    Slug and order_index are fixed — only thresholds and display fields are editable.
    """
    result = await db.execute(select(Level).where(Level.id == level_id))
    level = result.scalar_one_or_none()
    if not level:
        raise NotFoundException("Nível não encontrado")

    update_data = data.model_dump(exclude_unset=True)
    if not update_data:
        raise BadRequestException("Nenhum campo para atualizar")

    for field, value in update_data.items():
        setattr(level, field, value)

    await log_audit(
        db, admin_id=current_admin.id, action="update_level",
        entity_type="level", entity_id=str(level_id),
        details=update_data,
        ip_address=request.client.host if request.client else None,
    )

    await db.flush()

    # If requires_approval was turned OFF, auto-promote eligible users
    if "requires_approval" in update_data and not level.requires_approval:
        from sqlalchemy.orm import selectinload
        eligible = await db.execute(
            select(Profile).where(
                Profile.total_points >= level.min_points,
                Profile.current_level_id != level.id,
            ).options(selectinload(Profile.current_level))
        )
        for profile in eligible.scalars().all():
            current_order = profile.current_level.order_index if profile.current_level else 0
            if level.order_index > current_order:
                profile.current_level_id = level.id
                profile.level_pending_approval = False
        await db.flush()

    return LevelResponse.model_validate(level)


# ═══ USERS MANAGEMENT ═══
@router.get("/users")
async def list_users(
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    role: str | None = None,
    level_id: uuid.UUID | None = None,
    region_id: uuid.UUID | None = None,
    search: str | None = None,
):
    """Admin: List all users with filters."""
    service = UserService(db)
    params = PaginationParams(page=page, page_size=page_size)
    result = await service.list_users(
        params=params, role=role, level_id=level_id, region_id=region_id, search=search
    )
    # Serialize SQLAlchemy Profile objects to Pydantic
    result.items = [ProfileResponse.model_validate(u) for u in result.items]
    return result


@router.get("/users/{user_id}")
async def get_user_detail(
    user_id: uuid.UUID,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Admin: Get complete user details with activities, missions, badges."""
    from sqlalchemy.orm import selectinload
    from src.modules.gamification.models import UserBadge

    # Profile with level
    result = await db.execute(
        select(Profile)
        .options(selectinload(Profile.current_level))
        .where(Profile.id == user_id)
    )
    user = result.scalar_one_or_none()
    if not user:
        raise NotFoundException("Usuário não encontrado")

    # Recent activities (last 20)
    activities_result = await db.execute(
        select(Activity)
        .where(Activity.user_id == user_id)
        .order_by(Activity.created_at.desc())
        .limit(20)
    )
    activities = [
        {
            "id": str(a.id),
            "action_type": a.action_type,
            "action_description": a.action_description,
            "points_awarded": a.points_awarded,
            "created_at": a.created_at.isoformat() if a.created_at else None,
        }
        for a in activities_result.scalars().all()
    ]

    # User missions
    missions_result = await db.execute(
        select(UserMission)
        .options(selectinload(UserMission.mission))
        .where(UserMission.user_id == user_id)
        .order_by(UserMission.started_at.desc())
    )
    missions = [
        {
            "id": str(um.id),
            "mission_title": um.mission.title if um.mission else "—",
            "status": um.status,
            "created_at": um.started_at.isoformat() if um.started_at else None,
            "completed_at": um.completed_at.isoformat() if um.completed_at else None,
        }
        for um in missions_result.scalars().all()
    ]

    # Badges
    badges_result = await db.execute(
        select(UserBadge)
        .options(selectinload(UserBadge.badge))
        .where(UserBadge.user_id == user_id)
        .order_by(UserBadge.awarded_at.desc())
    )
    badges = [
        {
            "id": str(ub.id),
            "name": ub.badge.name if ub.badge else "—",
            "description": ub.badge.description if ub.badge else "",
            "icon_url": ub.badge.icon_url if ub.badge else None,
            "awarded_at": ub.awarded_at.isoformat() if ub.awarded_at else None,
        }
        for ub in badges_result.scalars().all()
    ]

    # Event participation count
    event_count = (await db.execute(
        select(func.count(EventParticipant.id)).where(EventParticipant.user_id == user_id)
    )).scalar() or 0

    checkin_count = (await db.execute(
        select(func.count(EventParticipant.id)).where(
            EventParticipant.user_id == user_id,
            EventParticipant.check_in_at.isnot(None),
        )
    )).scalar() or 0

    # Referred by
    referred_by_name = None
    if user.referred_by:
        ref = await db.execute(select(Profile.full_name).where(Profile.id == user.referred_by))
        referred_by_name = ref.scalar_one_or_none()

    # Referrals count
    referral_count = (await db.execute(
        select(func.count(Profile.id)).where(Profile.referred_by == user_id)
    )).scalar() or 0

    return {
        "profile": {
            "id": str(user.id),
            "full_name": user.full_name,
            "email": user.email,
            "phone": user.phone,
            "cpf": user.cpf,
            "bio": user.bio,
            "avatar_url": user.avatar_url,
            "birth_date": user.birth_date.isoformat() if user.birth_date else None,
            "gender": user.gender,
            "neighborhood": user.neighborhood,
            "city": user.city,
            "state": user.state,
            "zip_code": user.zip_code,
            "total_points": user.total_points,
            "role": user.role,
            "referral_code": user.referral_code,
            "referred_by_name": referred_by_name,
            "referral_count": referral_count,
            "onboarding_completed": user.onboarding_completed,
            "level_pending_approval": user.level_pending_approval,
            "is_active": user.is_active,
            "created_at": user.created_at.isoformat() if user.created_at else None,
            "current_level": {
                "name": user.current_level.name,
                "color": user.current_level.color,
            } if user.current_level else None,
        },
        "activities": activities,
        "missions": missions,
        "badges": badges,
        "events": {
            "registered": event_count,
            "checked_in": checkin_count,
        },
    }

@router.patch("/users/{user_id}/role")
async def update_user_role(
    user_id: uuid.UUID,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    request: Request,
    role: str = Query(...),
):
    """Admin: Update a user's role. Requires SUPER_ADMIN or CAMPAIGN_MANAGER."""
    valid_roles = {"participant", "moderator", "regional_coordinator", "campaign_manager", "analyst", "super_admin"}
    if role not in valid_roles:
        raise BadRequestException(f"Role inválido. Opções: {', '.join(valid_roles)}")

    await db.execute(
        sql_update(Profile).where(Profile.id == user_id).values(role=role)
    )

    # Audit log (PRD §7.2)
    await log_audit(
        db, admin_id=current_admin.id, action="update_role",
        entity_type="user", entity_id=str(user_id),
        details={"new_role": role},
        ip_address=request.client.host if request.client else None,
    )

    return {"message": f"Role atualizado para {role}"}


@router.post("/users/{user_id}/approve-level")
async def approve_user_level(
    user_id: uuid.UUID,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    request: Request,
    approved: bool = Query(default=True),
):
    """
    Approve or reject a level-up request (PRD §3.2).
    Levels 4 and 5 require human approval.
    """
    profile_result = await db.execute(select(Profile).where(Profile.id == user_id))
    profile = profile_result.scalar_one_or_none()
    if not profile:
        raise NotFoundException("Usuário não encontrado")

    if not profile.level_pending_approval:
        raise BadRequestException("Usuário não possui aprovação de nível pendente")

    if approved:
        # Find the next level that requires approval
        current_order = 0
        if profile.current_level_id:
            level_result = await db.execute(select(Level).where(Level.id == profile.current_level_id))
            current_level = level_result.scalar_one_or_none()
            current_order = current_level.order_index if current_level else 0

        next_level_result = await db.execute(
            select(Level).where(
                Level.order_index > current_order,
                Level.requires_approval.is_(True),
            ).order_by(Level.order_index.asc()).limit(1)
        )
        next_level = next_level_result.scalar_one_or_none()

        if next_level:
            await db.execute(
                sql_update(Profile).where(Profile.id == user_id).values(
                    current_level_id=next_level.id,
                    level_pending_approval=False,
                )
            )

            # Notify user
            notification = Notification(
                user_id=user_id,
                title="🎉 Nível aprovado!",
                body=f"Parabéns! Seu nível foi aprovado: {next_level.name}",
                notification_type="level_up",
            )
            db.add(notification)
    else:
        await db.execute(
            sql_update(Profile).where(Profile.id == user_id).values(
                level_pending_approval=False,
            )
        )

    # Audit log
    await log_audit(
        db, admin_id=current_admin.id, action="approve_level" if approved else "reject_level",
        entity_type="user", entity_id=str(user_id),
        details={"approved": approved},
        ip_address=request.client.host if request.client else None,
    )

    return {"message": "Nível aprovado" if approved else "Solicitação de nível negada"}


@router.post("/users/{user_id}/suspend")
async def suspend_user(
    user_id: uuid.UUID,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    request: Request,
    reason: str = Query(default=""),
):
    """Admin: Suspend a user account."""
    await db.execute(
        sql_update(Profile).where(Profile.id == user_id).values(is_active=False)
    )

    await log_audit(
        db, admin_id=current_admin.id, action="suspend_user",
        entity_type="user", entity_id=str(user_id),
        details={"reason": reason},
        ip_address=request.client.host if request.client else None,
    )

    return {"message": "Usuário suspenso"}


# ═══ MISSIONS MANAGEMENT ═══
@router.get("/missions")
async def list_missions(
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    category_id: uuid.UUID | None = None,
    mission_type: str | None = None,
    is_featured: bool | None = None,
):
    """Admin: List all missions (including inactive)."""
    service = MissionService(db)
    params = PaginationParams(page=page, page_size=page_size)
    result = await service.list_missions(
        params=params,
        category_id=category_id,
        mission_type=mission_type,
        is_featured=is_featured,
        include_inactive=True,
    )
    result.items = [MissionResponse.model_validate(m) for m in result.items]
    return result


@router.post("/missions")
async def create_mission(
    data: MissionCreate,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    request: Request,
):
    """Admin: Create a new mission."""
    service = MissionService(db)
    mission = await service.create_mission(data)

    await log_audit(
        db, admin_id=current_admin.id, action="create_mission",
        entity_type="mission", entity_id=str(mission.id),
        ip_address=request.client.host if request.client else None,
    )

    return {"message": "Missão criada", "mission_id": str(mission.id)}


@router.patch("/missions/{mission_id}")
async def update_mission(
    mission_id: uuid.UUID,
    data: MissionUpdate,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    request: Request,
):
    """Admin: Update a mission."""
    service = MissionService(db)
    await service.update_mission(mission_id, data)

    await log_audit(
        db, admin_id=current_admin.id, action="update_mission",
        entity_type="mission", entity_id=str(mission_id),
        details=data.model_dump(exclude_unset=True),
        ip_address=request.client.host if request.client else None,
    )

    return {"message": "Missão atualizada"}


@router.delete("/missions/{mission_id}")
async def delete_mission(
    mission_id: uuid.UUID,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    request: Request,
):
    """Admin: Soft delete a mission."""
    service = MissionService(db)
    await service.delete_mission(mission_id)

    await log_audit(
        db, admin_id=current_admin.id, action="delete_mission",
        entity_type="mission", entity_id=str(mission_id),
        ip_address=request.client.host if request.client else None,
    )

    return {"message": "Missão desativada com sucesso"}


@router.post("/missions/{mission_id}/verify")
async def verify_mission(
    mission_id: uuid.UUID,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    request: Request,
    user_id: uuid.UUID = Query(...),
    approved: bool = Query(default=True),
    rejection_reason: str | None = Query(default=None),
):
    """Admin: Verify a mission submission (PRD §4.2 VALIDATED/REJECTED)."""
    result = await db.execute(
        select(UserMission)
        .where(UserMission.user_id == user_id, UserMission.mission_id == mission_id)
    )
    user_mission = result.scalar_one_or_none()
    if not user_mission:
        raise NotFoundException("Submissão não encontrada")

    if user_mission.status != "submitted":
        raise BadRequestException(f"Missão não está em estado de submissão (atual: {user_mission.status})")

    if approved:
        user_mission.status = "completed"
        user_mission.completed_at = datetime.now(timezone.utc)
        user_mission.verified_at = datetime.now(timezone.utc)
        user_mission.verified_by = current_admin.id

        # Get mission for points
        mission = await db.execute(select(Mission).where(Mission.id == mission_id))
        mission_obj = mission.scalar_one()
        user_mission.points_awarded = mission_obj.points_reward

        engine = GamificationEngine(db)
        await engine.award_points(
            user_id=user_id,
            points=mission_obj.points_reward,
            action_type="mission_verified",
            description=f"Missão verificada: {mission_obj.title}",
            reference_type="mission",
            reference_id=mission_id,
            idempotency_key=f"{user_id}:mission_verified:{mission_id}:{user_mission.id}",
        )
    else:
        user_mission.status = "rejected"
        user_mission.rejected_reason = rejection_reason

    await log_audit(
        db, admin_id=current_admin.id,
        action="verify_mission" if approved else "reject_mission",
        entity_type="user_mission", entity_id=str(user_mission.id),
        details={"user_id": str(user_id), "approved": approved, "reason": rejection_reason},
        ip_address=request.client.host if request.client else None,
    )

    return {"message": "Missão verificada" if approved else "Missão rejeitada"}


# ═══ BADGES MANAGEMENT ═══

@router.get("/badges", response_model=list[BadgeResponse])
async def list_badges(
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Admin: List all badges."""
    result = await db.execute(select(Badge).order_by(Badge.name))
    return list(result.scalars().all())


@router.post("/badges")
async def create_badge(
    data: BadgeCreate,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    request: Request,
):
    """Admin: Create a new badge."""
    badge = Badge(**data.model_dump())
    db.add(badge)
    await db.flush()

    await log_audit(
        db, admin_id=current_admin.id, action="create_badge",
        entity_type="badge", entity_id=str(badge.id),
        ip_address=request.client.host if request.client else None,
    )

    return {"message": "Badge criado", "badge_id": str(badge.id)}


@router.patch("/badges/{badge_id}")
async def update_badge(
    badge_id: uuid.UUID,
    data: BadgeUpdate,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    request: Request,
):
    """Admin: Update an existing badge."""
    from fastapi import HTTPException
    result = await db.execute(select(Badge).where(Badge.id == badge_id))
    badge = result.scalar_one_or_none()
    if not badge:
        raise HTTPException(status_code=404, detail="Badge não encontrado")

    update_data = data.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(badge, key, value)
    
    await db.flush()

    await log_audit(
        db, admin_id=current_admin.id, action="update_badge",
        entity_type="badge", entity_id=str(badge.id),
        ip_address=request.client.host if request.client else None,
    )

    return {"message": "Badge atualizado"}


@router.delete("/badges/{badge_id}")
async def delete_badge(
    badge_id: uuid.UUID,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    request: Request,
):
    """Admin: Soft delete / deactivate a badge."""
    from fastapi import HTTPException
    result = await db.execute(select(Badge).where(Badge.id == badge_id))
    badge = result.scalar_one_or_none()
    if not badge:
        raise HTTPException(status_code=404, detail="Badge não encontrado")

    badge.is_active = False
    await db.flush()

    await log_audit(
        db, admin_id=current_admin.id, action="delete_badge",
        entity_type="badge", entity_id=str(badge.id),
        ip_address=request.client.host if request.client else None,
    )

    return {"message": "Badge inativado"}


# ═══ MODERATION QUEUE (PRD §7.1.8) ═══
@router.get("/moderation/queue")
async def get_moderation_queue(
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
):
    """
    Get the moderation queue — missions pending verification (PRD §7.1.8).
    Also includes anomaly detection for suspicious point spikes (PRD §4.3).
    """
    from sqlalchemy.orm import selectinload

    # Pending verifications
    query = (
        select(UserMission)
        .options(
            selectinload(UserMission.mission),
            selectinload(UserMission.user),
        )
        .where(UserMission.status == "submitted")
        .order_by(UserMission.submitted_at.asc())
    )
    params = PaginationParams(page=page, page_size=page_size)
    count_query = select(func.count(UserMission.id)).where(UserMission.status == "submitted")
    total = (await db.execute(count_query)).scalar() or 0

    query = query.offset(params.offset).limit(params.page_size)
    result = await db.execute(query)
    items = list(result.scalars().all())

    # PRD §4.3: Anomaly detection — users with suspicious point spikes in last 24h
    day_ago = datetime.now(timezone.utc) - timedelta(hours=24)
    suspicious = await db.execute(
        select(
            Activity.user_id,
            func.sum(Activity.points_awarded).label("total_24h"),
            func.count(Activity.id).label("action_count"),
        )
        .where(Activity.created_at >= day_ago)
        .group_by(Activity.user_id)
        .having(func.sum(Activity.points_awarded) > 500)  # Threshold
        .order_by(func.sum(Activity.points_awarded).desc())
        .limit(20)
    )

    suspicious_users = [
        {"user_id": row.user_id, "points_24h": row.total_24h, "actions_24h": row.action_count}
        for row in suspicious
    ]

    return {
        "pending_verifications": {
            "items": items,
            "total": total,
            "page": params.page,
            "page_size": params.page_size,
        },
        "suspicious_activity": suspicious_users,
    }


# ═══ EVENTS MANAGEMENT ═══

@router.get("/events")
async def list_events_admin(
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=100),
    upcoming_only: bool = False,
):
    """Admin: List all events (including past and inactive)."""
    service = EventService(db)
    params = PaginationParams(page=page, page_size=page_size)
    result = await service.list_events(
        params=params, upcoming_only=upcoming_only, include_inactive=True,
    )
    items = []
    for e in result.items:
        data = EventResponse.model_validate(e)
        data.participants_count = len(e.participants) if e.participants else 0
        items.append(data)
    return {
        "items": items,
        "total": result.total,
        "page": result.page,
        "page_size": result.page_size,
        "total_pages": result.total_pages,
    }


@router.post("/events")
async def create_event(
    data: EventCreate,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    request: Request,
):
    """Admin: Create a new event."""
    service = EventService(db)
    event = await service.create_event(data, created_by=None)

    await log_audit(
        db, admin_id=current_admin.id, action="create_event",
        entity_type="event", entity_id=str(event.id),
        ip_address=request.client.host if request.client else None,
    )

    return {
        "message": "Evento criado",
        "event_id": str(event.id),
        "checkin_code": event.checkin_code,
    }


@router.patch("/events/{event_id}")
async def update_event(
    event_id: uuid.UUID,
    data: EventUpdate,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    request: Request,
):
    """Admin: Update an event."""
    service = EventService(db)
    await service.update_event(event_id, data)

    await log_audit(
        db, admin_id=current_admin.id, action="update_event",
        entity_type="event", entity_id=str(event_id),
        ip_address=request.client.host if request.client else None,
    )

    return {"message": "Evento atualizado"}


@router.post("/events/{event_id}/regenerate-code")
async def regenerate_checkin_code(
    event_id: uuid.UUID,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Admin: Regenerate check-in code for an event."""
    service = EventService(db)
    new_code = await service.regenerate_checkin_code(event_id)
    return {"message": "Código regenerado", "checkin_code": new_code}


@router.get("/events/{event_id}/participants")
async def list_event_participants(
    event_id: uuid.UUID,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Admin: List participants of an event with user details and status summary."""
    # Get event
    event_result = await db.execute(
        select(Event).where(Event.id == event_id)
    )
    event = event_result.scalar_one_or_none()
    if not event:
        raise HTTPException(status_code=404, detail="Evento não encontrado")

    # Get participants with user profiles
    from src.modules.users.models import Profile

    result = await db.execute(
        select(EventParticipant, Profile)
        .join(Profile, EventParticipant.user_id == Profile.id)
        .where(EventParticipant.event_id == event_id)
        .order_by(EventParticipant.registered_at.desc())
    )
    rows = result.all()

    # Build status summary
    status_counts: dict[str, int] = {}
    participants = []
    for participant, profile in rows:
        status = participant.status or "registered"
        status_counts[status] = status_counts.get(status, 0) + 1
        participants.append({
            "id": str(participant.id),
            "user_id": str(profile.id),
            "full_name": profile.full_name,
            "email": profile.email,
            "phone": profile.phone,
            "avatar_url": profile.avatar_url,
            "status": status,
            "registered_at": participant.registered_at.isoformat() if participant.registered_at else None,
            "check_in_at": participant.check_in_at.isoformat() if participant.check_in_at else None,
        })

    return {
        "event": {
            "id": str(event.id),
            "title": event.title,
            "event_type": event.event_type,
            "start_datetime": event.start_datetime.isoformat(),
            "end_datetime": event.end_datetime.isoformat() if event.end_datetime else None,
            "location_name": event.location_name,
            "max_capacity": event.max_capacity,
            "points_reward": event.points_reward,
        },
        "summary": {
            "total": len(participants),
            "by_status": status_counts,
        },
        "participants": participants,
    }

# ═══ CONTENT MANAGEMENT ═══
@router.get("/content")
async def list_admin_content(
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=100),
):
    """Admin: List all content (including inactive)."""
    from src.modules.content.schemas import ContentResponse
    from src.shared.pagination import PaginatedResponse, PaginationParams
    
    query = select(Content).order_by(Content.created_at.desc())
    count_query = select(func.count(Content.id))
    
    total = (await db.execute(count_query)).scalar() or 0
    params = PaginationParams(page=page, page_size=page_size)
    query = query.offset(params.offset).limit(params.page_size)
    
    result = await db.execute(query)
    items = list(result.scalars().all())
    
    return {
        "items": [ContentResponse.model_validate(c) for c in items],
        "total": total,
        "page": params.page,
        "page_size": params.page_size,
    }


@router.delete("/content/{content_id}")
async def delete_content(
    content_id: uuid.UUID,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    request: Request,
):
    """Admin: Soft delete content."""
    from fastapi import HTTPException
    result = await db.execute(select(Content).where(Content.id == content_id))
    content = result.scalar_one_or_none()
    if not content:
        raise HTTPException(status_code=404, detail="Conteúdo não encontrado")

    content.is_active = False
    await db.flush()

    await log_audit(
        db, admin_id=current_admin.id, action="delete_content",
        entity_type="content", entity_id=str(content.id),
        ip_address=request.client.host if request.client else None,
    )

    return {"message": "Conteúdo inativado"}


@router.post("/content")
async def create_content(
    data: ContentCreate,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    request: Request,
):
    """Admin: Create new content for sharing."""
    service = ContentService(db)
    content = await service.create_content(data, created_by=None)

    await log_audit(
        db, admin_id=current_admin.id, action="create_content",
        entity_type="content", entity_id=str(content.id),
        ip_address=request.client.host if request.client else None,
    )

    return {"message": "Conteúdo criado", "content_id": str(content.id)}


@router.patch("/content/{content_id}")
async def update_content(
    content_id: uuid.UUID,
    data: ContentUpdate,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    request: Request,
):
    """Admin: Update content."""
    service = ContentService(db)
    await service.update_content(content_id, data)

    await log_audit(
        db, admin_id=current_admin.id, action="update_content",
        entity_type="content", entity_id=str(content_id),
        ip_address=request.client.host if request.client else None,
    )

    return {"message": "Conteúdo atualizado"}


# ═══ NOTIFICATIONS & CAMPAIGNS ═══

class SendNotificationRequest(BaseModel):
    title: str
    body: str
    notification_type: str = "info"
    user_id: uuid.UUID | None = None
    target_level_id: uuid.UUID | None = None
    target_region_id: uuid.UUID | None = None


@router.get("/notifications")
async def list_notifications(
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
):
    """Admin: List sent notifications grouped by broadcast batch."""
    from sqlalchemy import String as SAString, func as sa_func

    # Only show admin-sent notification types (exclude system-generated like level_up, level_approval)
    admin_types = ("info", "campaign", "event", "mission", "system")

    # Get distinct notification batches (grouped by title + body + type + sent_at truncated to second)
    subq = (
        select(
            sa_func.min(sa_func.cast(Notification.id, SAString)).label("id"),
            Notification.title,
            Notification.body,
            Notification.notification_type,
            sa_func.min(Notification.sent_at).label("sent_at"),
            sa_func.count(Notification.id).label("sent_count"),
            sa_func.sum(
                sa_func.cast(Notification.is_read, Integer)
            ).label("read_count"),
        )
        .where(Notification.notification_type.in_(admin_types))
        .group_by(
            Notification.title,
            Notification.body,
            Notification.notification_type,
            sa_func.date_trunc("minute", Notification.sent_at),
        )
        .order_by(sa_func.min(Notification.sent_at).desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
    )

    result = await db.execute(subq)
    rows = result.all()

    # Count total groups
    count_q = (
        select(sa_func.count())
        .select_from(
            select(Notification.title)
            .where(Notification.notification_type.in_(admin_types))
            .group_by(
                Notification.title,
                Notification.body,
                Notification.notification_type,
                sa_func.date_trunc("minute", Notification.sent_at),
            )
            .subquery()
        )
    )
    total = (await db.execute(count_q)).scalar() or 0

    items = []
    for row in rows:
        items.append({
            "id": str(row.id),
            "title": row.title,
            "body": row.body,
            "notification_type": row.notification_type,
            "sent_at": row.sent_at.isoformat() if row.sent_at else None,
            "sent_count": row.sent_count,
            "read_count": row.read_count or 0,
        })

    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
    }

@router.post("/notifications/send")
async def send_notification(
    data: SendNotificationRequest,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    request: Request,
):
    """Admin: Send a notification (targeted or broadcast)."""
    service = NotificationService(db)

    if data.user_id:
        # Send to specific user
        await service.create_notification(
            title=data.title, body=data.body,
            notification_type=data.notification_type,
            user_id=data.user_id,
            target_level_id=data.target_level_id,
            target_region_id=data.target_region_id,
        )
        sent_count = 1
    else:
        # Broadcast: create notification for ALL active users
        result = await db.execute(
            select(Profile.id).where(Profile.is_active.is_(True))
        )
        user_ids = [row[0] for row in result.all()]
        for uid in user_ids:
            notif = Notification(
                user_id=uid,
                title=data.title,
                body=data.body,
                notification_type=data.notification_type,
            )
            db.add(notif)
        await db.flush()
        sent_count = len(user_ids)

    await log_audit(
        db, admin_id=current_admin.id, action="send_notification",
        entity_type="notification", entity_id="broadcast",
        details={"title": data.title, "sent_count": sent_count},
        ip_address=request.client.host if request.client else None,
    )

    # ═══ Push Notifications ═══
    # Dispara push para os dispositivos registrados dos destinatários.
    # Falha no push NÃO impede a notificação in-app (já criada acima).
    push_sent = 0
    try:
        from src.modules.push.service import PushService
        push_service = PushService(db)
        if data.user_id:
            await push_service.send_to_user(data.user_id, data.title, data.body)
        else:
            await push_service.send_to_users(user_ids, data.title, data.body)
        push_sent = sent_count
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning("Push notification failed: %s", e)

    return {
        "message": f"Notificação enviada para {sent_count} usuário(s)",
        "sent_count": sent_count,
        "push_sent": push_sent,
    }


@router.post("/notifications/campaign")
async def create_campaign(
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    request: Request,
    title: str = Query(...),
    body: str = Query(...),
    segment_level_id: uuid.UUID | None = None,
    segment_region_id: uuid.UUID | None = None,
    segment_role: str | None = None,
):
    """
    Admin: Create and send a segmented notification campaign (PRD §7.1.3).
    // LEGAL-REVIEW: Comunicação segmentada requer consentimento de comunicação (PRD §8.2)
    """
    service = NotificationService(db)
    campaign = await service.create_campaign(
        title=title, body=body, created_by=current_admin.id,
        segment_level_id=segment_level_id,
        segment_region_id=segment_region_id,
        segment_role=segment_role,
    )
    result = await service.send_campaign(campaign.id)

    await log_audit(
        db, admin_id=current_admin.id, action="send_campaign",
        entity_type="notification_campaign", entity_id=str(campaign.id),
        details={"sent_count": result.get("sent_count", 0)},
        ip_address=request.client.host if request.client else None,
    )

    return result


@router.get("/notifications/campaigns")
async def list_campaigns(
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Admin: List all notification campaigns."""
    service = NotificationService(db)
    return await service.list_campaigns()


# ═══ SYSTEM NOTIFICATION CONFIGS ═══
@router.get("/notifications/system-configs")
async def list_system_notification_configs(
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Admin: List all system notification configs with active/inactive status."""
    service = SystemNotificationService(db)
    configs = await service.get_all_configs()
    return [
        {
            "event_key": c.event_key,
            "label": c.label,
            "description": c.description,
            "title_template": c.title_template,
            "body_template": c.body_template,
            "notification_type": c.notification_type,
            "is_active": c.is_active,
        }
        for c in configs
    ]


@router.patch("/notifications/system-configs/{event_key}")
async def toggle_system_notification_config(
    event_key: str,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    request: Request,
    is_active: bool = Query(...),
):
    """Admin: Toggle a system notification config on/off."""
    service = SystemNotificationService(db)
    config = await service.toggle_config(event_key, is_active)
    if not config:
        raise NotFoundException(f"Config '{event_key}' não encontrada")

    await log_audit(
        db, admin_id=current_admin.id,
        action="toggle_system_notification",
        entity_type="system_notification_config",
        entity_id=event_key,
        details={"is_active": is_active},
        ip_address=request.client.host if request.client else None,
    )

    return {
        "event_key": config.event_key,
        "label": config.label,
        "is_active": config.is_active,
        "message": f"Notificação '{config.label}' {'ativada' if is_active else 'desativada'}",
    }


# ═══ ANALYTICS ═══
@router.get("/analytics/engagement")
async def get_engagement_analytics(
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    days: int = Query(default=30, ge=1, le=365),
):
    """Admin: Get engagement analytics for the specified period."""
    start_date = datetime.now(timezone.utc) - timedelta(days=days)

    # Activities per day
    daily_activities = await db.execute(
        select(
            func.date_trunc("day", Activity.created_at).label("date"),
            func.count(Activity.id).label("count"),
            func.sum(Activity.points_awarded).label("points"),
        )
        .where(Activity.created_at >= start_date)
        .group_by(func.date_trunc("day", Activity.created_at))
        .order_by(func.date_trunc("day", Activity.created_at))
    )

    # New registrations per day
    daily_registrations = await db.execute(
        select(
            func.date_trunc("day", Profile.created_at).label("date"),
            func.count(Profile.id).label("count"),
        )
        .where(Profile.created_at >= start_date)
        .group_by(func.date_trunc("day", Profile.created_at))
        .order_by(func.date_trunc("day", Profile.created_at))
    )

    # Missions completed per day
    daily_missions = await db.execute(
        select(
            func.date_trunc("day", UserMission.completed_at).label("date"),
            func.count(UserMission.id).label("count"),
        )
        .where(UserMission.completed_at >= start_date, UserMission.status == "completed")
        .group_by(func.date_trunc("day", UserMission.completed_at))
        .order_by(func.date_trunc("day", UserMission.completed_at))
    )

    return {
        "period_days": days,
        "daily_activities": [
            {"date": str(row.date), "count": row.count, "points": row.points}
            for row in daily_activities
        ],
        "daily_registrations": [
            {"date": str(row.date), "count": row.count}
            for row in daily_registrations
        ],
        "daily_missions_completed": [
            {"date": str(row.date), "count": row.count}
            for row in daily_missions
        ],
    }


# ═══ REPORTS EXPORT (PRD §7.1.7) ═══
@router.get("/reports/export")
async def export_report(
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    report_type: str = Query(default="users", description="users, missions, events"),
    days: int = Query(default=30, ge=1, le=365),
):
    """Admin: Export report as CSV (PRD §7.1.7)."""
    start_date = datetime.now(timezone.utc) - timedelta(days=days)
    output = io.StringIO()
    writer = csv.writer(output)

    if report_type == "users":
        writer.writerow(["ID", "Nome", "Email", "Cidade", "Estado", "Pontos", "Nível", "Role", "Cadastro"])
        result = await db.execute(
            select(Profile)
            .where(Profile.is_active.is_(True), Profile.created_at >= start_date)
            .order_by(Profile.created_at.desc())
        )
        for user in result.scalars().all():
            writer.writerow([
                str(user.id), user.full_name, user.email,
                user.city, user.state, user.total_points,
                "", user.role, str(user.created_at),
            ])

    elif report_type == "missions":
        writer.writerow(["Usuário", "Missão", "Status", "Pontos", "Início", "Conclusão"])
        from sqlalchemy.orm import selectinload
        result = await db.execute(
            select(UserMission)
            .options(selectinload(UserMission.mission), selectinload(UserMission.user))
            .where(UserMission.started_at >= start_date)
            .order_by(UserMission.started_at.desc())
        )
        for um in result.scalars().all():
            writer.writerow([
                um.user.full_name if um.user else "", um.mission.title if um.mission else "",
                um.status, um.points_awarded, str(um.started_at),
                str(um.completed_at) if um.completed_at else "",
            ])

    elif report_type == "events":
        writer.writerow(["Evento", "Tipo", "Data", "Local", "Inscritos", "Check-ins"])
        result = await db.execute(
            select(Event).where(Event.start_datetime >= start_date).order_by(Event.start_datetime.desc())
        )
        for event in result.scalars().all():
            registered = (await db.execute(
                select(func.count(EventParticipant.id)).where(EventParticipant.event_id == event.id)
            )).scalar() or 0
            attended = (await db.execute(
                select(func.count(EventParticipant.id)).where(
                    EventParticipant.event_id == event.id, EventParticipant.status == "attended"
                )
            )).scalar() or 0
            writer.writerow([
                event.title, event.event_type, str(event.start_datetime),
                event.location_name or "", registered, attended,
            ])

    output.seek(0)
    return StreamingResponse(
        io.BytesIO(output.getvalue().encode("utf-8-sig")),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=relatorio_{report_type}_{days}d.csv"},
    )


# ═══ AUDIT LOGS (PRD §7.2) ═══
@router.get("/audit-logs")
async def get_audit_logs(
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=100),
    action: str | None = None,
    entity_type: str | None = None,
):
    """Admin: View audit logs. SUPER_ADMIN only."""
    query = select(AuditLog)
    count_query = select(func.count(AuditLog.id))

    if action:
        query = query.where(AuditLog.action == action)
        count_query = count_query.where(AuditLog.action == action)
    if entity_type:
        query = query.where(AuditLog.entity_type == entity_type)
        count_query = count_query.where(AuditLog.entity_type == entity_type)

    total = (await db.execute(count_query)).scalar() or 0
    params = PaginationParams(page=page, page_size=page_size)
    query = query.order_by(AuditLog.created_at.desc()).offset(params.offset).limit(params.page_size)

    result = await db.execute(query)
    items = list(result.scalars().all())

    from src.shared.pagination import PaginatedResponse
    # Serialize AuditLog SQLAlchemy objects
    serialized = [
        {
            "id": str(log.id),
            "admin_id": str(log.admin_id),
            "action": log.action,
            "entity_type": log.entity_type,
            "entity_id": log.entity_id,
            "details": log.details,
            "ip_address": log.ip_address,
            "created_at": log.created_at.isoformat() if log.created_at else None,
        }
        for log in items
    ]
    return PaginatedResponse.create(items=serialized, total=total, params=params)


# ═══ POINTS RECONCILIATION (PRD §5.1) ═══
@router.post("/reconcile-points/{user_id}")
async def reconcile_user_points(
    user_id: uuid.UUID,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Admin: Reconcile materialized points vs ledger for a user."""
    engine = GamificationEngine(db)
    return await engine.reconcile_points(user_id)


# ═══ POINT CONFIGS ═══

class PointConfigResponse(BaseModel):
    id: uuid.UUID
    key: str
    points: int
    label: str
    description: str | None = None
    category: str
    is_active: bool

    model_config = {"from_attributes": True}


class PointConfigUpdateRequest(BaseModel):
    points: int | None = None
    is_active: bool | None = None
    label: str | None = None
    description: str | None = None
    category: str | None = None


class PointConfigCreateRequest(BaseModel):
    key: str
    points: int
    label: str
    description: str | None = None
    category: str


@router.get("/point-configs", response_model=list[PointConfigResponse])
async def list_point_configs(
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Admin: List all point configurations grouped by category."""
    from src.modules.gamification.point_config import PointConfigService
    configs = await PointConfigService.get_all(db)
    return configs


@router.get("/point-configs/{key}", response_model=PointConfigResponse)
async def get_point_config(
    key: str,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Admin: Get a specific point configuration by key."""
    from src.modules.gamification.point_config import PointConfigService
    config = await PointConfigService.get_by_key(db, key)
    if not config:
        raise NotFoundException(f"Configuração '{key}' não encontrada")
    return config


@router.put("/point-configs/{key}", response_model=PointConfigResponse)
async def update_point_config(
    key: str,
    data: PointConfigUpdateRequest,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Admin: Update points value and/or active state for a configuration."""
    from src.modules.gamification.point_config import PointConfigService
    if data.points is not None and data.points < 0:
        raise BadRequestException("Pontos não podem ser negativos")
    config = await PointConfigService.get_by_key(db, key)
    if not config:
        raise NotFoundException(f"Configuração '{key}' não encontrada")

    changes: dict = {}
    if data.points is not None:
        config.points = data.points
        changes["new_points"] = data.points
    if data.is_active is not None:
        config.is_active = data.is_active
        changes["is_active"] = data.is_active
    if data.label is not None:
        config.label = data.label
        changes["label"] = data.label
    if data.description is not None:
        config.description = data.description
        changes["description"] = data.description
    if data.category is not None:
        config.category = data.category
        changes["category"] = data.category

    from datetime import datetime, timezone
    config.updated_at = datetime.now(timezone.utc)
    await db.flush()
    PointConfigService.invalidate_cache()

    await log_audit(
        db,
        admin_id=current_admin.id,
        action="point_config_update",
        entity_type="point_config",
        details={"key": key, **changes},
    )
    return config


@router.post("/point-configs", response_model=PointConfigResponse, status_code=201)
async def create_point_config(
    data: PointConfigCreateRequest,
    current_admin: Annotated[AdminUser, Depends(get_current_admin_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Admin: Create a new point configuration."""
    from src.modules.gamification.point_config import PointConfig, PointConfigService
    existing = await PointConfigService.get_by_key(db, data.key)
    if existing:
        raise BadRequestException(f"Configuração '{data.key}' já existe")
    config = PointConfig(
        key=data.key,
        points=data.points,
        label=data.label,
        description=data.description,
        category=data.category,
    )
    db.add(config)
    await db.flush()
    PointConfigService.invalidate_cache()
    await log_audit(
        db,
        admin_id=current_admin.id,
        action="point_config_create",
        entity_type="point_config",
        details={"key": data.key, "points": data.points},
    )
    return config
